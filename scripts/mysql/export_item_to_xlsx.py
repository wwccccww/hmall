#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
将 hm-item 数据库中的 item 表导出为 .xlsx。

依赖:
  pip install -r scripts/mysql/requirements-export.txt
  或: pip install pymysql openpyxl

示例:
  python scripts/mysql/export_item_to_xlsx.py
  python scripts/mysql/export_item_to_xlsx.py -o D:\\data\\items.xlsx
  set MYSQL_HOST=127.0.0.1&& set MYSQL_PASSWORD=123&& python scripts/mysql/export_item_to_xlsx.py

环境变量（可选，括号内为默认值）:
  MYSQL_HOST (127.0.0.1)
  MYSQL_PORT (3306)
  MYSQL_USER (root)
  MYSQL_PASSWORD (123)
  MYSQL_DATABASE (hm-item)
"""

from __future__ import annotations

import argparse
import os
import sys
from datetime import date, datetime
from decimal import Decimal


def _cell_value(v):
    if v is None:
        return None
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, (datetime, date)):
        return v
    if isinstance(v, (bytes, bytearray)):
        return v.decode("utf-8", errors="replace")
    return v


def main() -> int:
    parser = argparse.ArgumentParser(description="导出 hm-item.item 到 xlsx")
    parser.add_argument(
        "-o",
        "--output",
        default=None,
        help="输出文件路径，默认当前目录 item_export_YYYYMMDD_HHMMSS.xlsx",
    )
    parser.add_argument("--host", default=os.environ.get("MYSQL_HOST", "127.0.0.1"))
    parser.add_argument("--port", type=int, default=int(os.environ.get("MYSQL_PORT", "3306")))
    parser.add_argument("--user", default=os.environ.get("MYSQL_USER", "root"))
    parser.add_argument("--password", default=os.environ.get("MYSQL_PASSWORD", "123"))
    parser.add_argument("--database", default=os.environ.get("MYSQL_DATABASE", "hm-item"))
    parser.add_argument("--table", default="item", help="表名，默认 item")
    args = parser.parse_args()

    try:
        import pymysql
    except ImportError:
        print("请先安装: pip install pymysql openpyxl", file=sys.stderr)
        return 1

    try:
        from openpyxl import Workbook
    except ImportError:
        print("请先安装: pip install openpyxl", file=sys.stderr)
        return 1

    out_path = args.output
    if not out_path:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = os.path.join(os.getcwd(), f"item_export_{ts}.xlsx")

    conn = pymysql.connect(
        host=args.host,
        port=args.port,
        user=args.user,
        password=args.password,
        database=args.database,
        charset="utf8mb4",
        cursorclass=pymysql.cursors.Cursor,
    )
    try:
        with conn.cursor() as cur:
            cur.execute(f"SELECT * FROM `{args.table}`")
            rows = cur.fetchall()
            colnames = [d[0] for d in cur.description]
    finally:
        conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = args.table[:31]  # Excel sheet 名最长 31

    for c, name in enumerate(colnames, start=1):
        ws.cell(row=1, column=c, value=name)

    for r, row in enumerate(rows, start=2):
        for c, v in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=_cell_value(v))

    wb.save(out_path)
    print(f"已写入: {out_path}  （共 {len(rows)} 行）")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
